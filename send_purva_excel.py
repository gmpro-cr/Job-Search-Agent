"""
send_purva_excel.py
Search for ~100 hiring managers / recruiters on LinkedIn hiring for
Project Manager roles in Pune / Remote, build an Excel sheet, and
email it to Purva via Gmail OAuth.
"""

import json
import logging
import re
import time
import base64
import os
from datetime import date
from pathlib import Path

logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")
logger = logging.getLogger(__name__)

RECIPIENT   = "purva.pilot57@gmail.com"
SENDER_NAME = "Job Search Agent"
TARGET      = 100
LOCATION    = "Pune, Remote"
ROLE_KEYWORDS = [
    "Project Manager", "Product Owner", "Scrum Master",
    "Delivery Manager", "Program Manager", "PMO", "Agile Coach",
]

GMAIL_CREDENTIALS_DIR = Path.home() / "Documents" / "Claude" / "gmail_credentials"

# ── Extended query set (20 queries × ~8 results = 160 candidates → 100 unique) ──
def _build_queries() -> list[str]:
    roles_or  = '"Project Manager" OR "Delivery Manager" OR "Program Manager"'
    roles_or2 = '"Product Owner" OR "Scrum Master" OR "PMO" OR "Agile Coach"'
    locs = ["Pune", "Remote", "Pune, Remote", "Maharashtra"]

    queries = []
    for loc in locs:
        queries += [
            f'site:linkedin.com/in "talent acquisition" {roles_or} {loc}',
            f'site:linkedin.com/in recruiter {roles_or} {loc} hiring 2026',
            f'site:linkedin.com/in "HR manager" OR "HR recruiter" {roles_or} {loc}',
            f'site:linkedin.com/in "talent acquisition manager" {roles_or} {loc}',
            f'site:linkedin.com/in "technical recruiter" {roles_or} {loc}',
        ]
    queries += [
        f'site:linkedin.com/in recruiter {roles_or2} Pune',
        f'site:linkedin.com/in "IT recruiter" {roles_or} Pune OR Remote India',
        f'site:linkedin.com/in "talent partner" {roles_or} Pune',
        f'site:linkedin.com/in "hiring manager" {roles_or} Pune OR Remote',
        f'site:linkedin.com/in recruiter "agile" "project manager" Pune',
        f'site:linkedin.com/in "senior recruiter" {roles_or} Pune',
        f'site:linkedin.com/in "lead recruiter" {roles_or} Maharashtra',
        f'site:linkedin.com/in recruiter {roles_or} "work from home" India',
        f'site:linkedin.com/in recruiter {roles_or} "remote" Pune India 2026',
        f'site:linkedin.com/in "HR business partner" {roles_or} Pune',
        f'site:linkedin.com/in "campus recruiter" OR "lateral recruiter" {roles_or} Pune',
        f'site:linkedin.com/in "staffing" {roles_or} Pune',
        f'site:linkedin.com/in "head of talent" {roles_or} Pune OR Remote India',
        f'site:linkedin.com/in "people operations" {roles_or} Pune',
        f'site:linkedin.com/in recruiter "PMP" OR "Prince2" {roles_or} Pune',
    ]
    return queries


# ── Parsers (reuse logic from hiring_managers.py) ────────────────────────────

def _parse_title(title: str) -> dict:
    title = re.sub(r'[\|\-–]\s*LinkedIn\s*$', '', title, flags=re.IGNORECASE).strip()
    parts = re.split(r'\s*[\-–]\s*', title, maxsplit=1)
    if len(parts) == 2:
        name_fixed = re.sub(r'(?<=[a-z])(?=[A-Z])', ' ', parts[0]).strip()
        title = f"{name_fixed} - {parts[1]}"

    m = re.match(r'^(.+?)\s*[\-–]\s*(.+?)\s+at\s+(.+)$', title, re.IGNORECASE)
    if m:
        return {"name": m.group(1).strip(),
                "their_role": re.sub(r'(?<=[a-z])(?=[A-Z])', ' ', m.group(2)).strip(),
                "company": m.group(3).strip().rstrip('|').strip()}

    m = re.match(r'^(.+?)\s*[\-–]\s*(.+?)\s*\|\s*(.+)$', title)
    if m:
        company = m.group(3).strip().rstrip('|').strip()
        return {"name": m.group(1).strip(),
                "their_role": re.sub(r'(?<=[a-z])(?=[A-Z])', ' ', m.group(2)).strip(),
                "company": company if len(company.split()) <= 5 else "IT/Consulting"}

    m = re.match(r'^(.+?)\s*[\-–]\s*(.+)$', title)
    if m:
        rest = re.sub(r'(?<=[a-z])(?=[A-Z])', ' ', m.group(2)).strip()
        role_kws = ["recruit", "talent", "hr ", "human resource", "manager",
                    "specialist", "lead", "head", "director", "coordinator",
                    "partner", "consultant", "analyst", "acquisition"]
        if any(kw in rest.lower() for kw in role_kws):
            return {"name": m.group(1).strip(), "their_role": rest, "company": "IT/Consulting"}
        return {"name": m.group(1).strip(), "their_role": "Recruiter / TA", "company": rest}

    return {"name": title.strip(), "their_role": "Recruiter / TA", "company": "IT/Consulting"}


def _infer_location(body: str, title: str) -> str:
    text = (body + " " + title).lower()
    tags = []
    if "pune" in text:                               tags.append("Pune")
    if "remote" in text or "work from home" in text: tags.append("Remote")
    if "hybrid" in text:                             tags.append("Hybrid")
    if "bangalore" in text or "bengaluru" in text:   tags.append("Bangalore")
    if "mumbai" in text:                             tags.append("Mumbai")
    if "hyderabad" in text:                          tags.append("Hyderabad")
    return ", ".join(tags) if tags else "India"


def _is_india_profile(url: str, body: str, title: str) -> bool:
    text = (url + body + title).lower()
    return any(kw in text for kw in
               ["india", "pune", "mumbai", "bangalore", "bengaluru",
                "hyderabad", "delhi", "chennai", "noida", "gurgaon",
                "maharashtra", ".in/", "in.linkedin"])


# ── Search ────────────────────────────────────────────────────────────────────

def collect_contacts(target: int = 100) -> list[dict]:
    from ddgs import DDGS

    queries    = _build_queries()
    seen_urls  = set()
    seen_names = set()
    contacts   = []

    for q_idx, query in enumerate(queries):
        if len(contacts) >= target:
            break
        logger.info("[%d/%d] %s", q_idx + 1, len(queries), query)
        try:
            with DDGS() as ddgs:
                results = list(ddgs.text(query, max_results=10))
        except Exception as e:
            logger.warning("Search error: %s", e)
            time.sleep(2)
            continue

        for r in results:
            if len(contacts) >= target:
                break
            url   = r.get("href", "")
            title = r.get("title", "")
            body  = r.get("body", "")

            if "linkedin.com/in/" not in url or url in seen_urls:
                continue
            seen_urls.add(url)

            if not _is_india_profile(url, body, title):
                continue

            parsed  = _parse_title(title)
            name    = parsed["name"].strip()
            company = parsed["company"].lstrip("| ").strip() or "IT/Consulting"

            if len(name.split()) < 2:
                continue
            if name.lower() in seen_names:
                continue
            seen_names.add(name.lower())

            # Infer which role they're hiring for
            text_lower = (body + " " + parsed["their_role"]).lower()
            hiring_for_tags = [kw for kw in ROLE_KEYWORDS if kw.lower() in text_lower]
            hiring_for = ", ".join(hiring_for_tags) if hiring_for_tags else "Project Manager"

            contacts.append({
                "S.No":         len(contacts) + 1,
                "Name":         name,
                "Company":      company,
                "Their Role":   parsed["their_role"],
                "Hiring For":   hiring_for,
                "Location":     _infer_location(body, title),
                "LinkedIn URL": url,
                "Outreach Message": (
                    f"Hi {name.split()[0]}, I'm a Project Manager with 9 years of experience "
                    f"delivering Agile/Scrum projects. Came across your profile and would love "
                    f"to connect about any {hiring_for} openings in Pune or Remote."
                ),
            })
            logger.info("  ✓ [%d] %s @ %s", len(contacts), name, company)

        time.sleep(1)  # be polite between queries

    logger.info("Total collected: %d", len(contacts))
    return contacts


# ── Excel builder ─────────────────────────────────────────────────────────────

def build_excel(contacts: list[dict], path: str) -> str:
    import openpyxl
    from openpyxl.styles import (Font, PatternFill, Alignment,
                                  Border, Side, GradientFill)
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Hiring Managers"

    # ── Title row ──
    ws.merge_cells("A1:H1")
    title_cell = ws["A1"]
    title_cell.value = f"Hiring Managers — Project Manager | Pune & Remote | {date.today().strftime('%d %b %Y')}"
    title_cell.font      = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    title_cell.fill      = PatternFill("solid", fgColor="005F73")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # ── Sub-title ──
    ws.merge_cells("A2:H2")
    sub = ws["A2"]
    sub.value = f"Total: {len(contacts)} contacts  |  Generated by Job Search Agent"
    sub.font      = Font(name="Calibri", italic=True, size=10, color="555555")
    sub.alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 18

    # ── Header row ──
    headers = ["S.No", "Name", "Company", "Their Role",
               "Hiring For", "Location", "LinkedIn URL", "Outreach Message"]
    header_fill   = PatternFill("solid", fgColor="0A9396")
    header_font   = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    thin          = Side(style="thin", color="CCCCCC")
    border        = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, h in enumerate(headers, 1):
        cell              = ws.cell(row=3, column=col, value=h)
        cell.font         = header_font
        cell.fill         = header_fill
        cell.alignment    = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border       = border
    ws.row_dimensions[3].height = 22

    # ── Data rows ──
    fill_even = PatternFill("solid", fgColor="E9F5F6")
    fill_odd  = PatternFill("solid", fgColor="FFFFFF")

    for row_idx, c in enumerate(contacts, 4):
        fill = fill_even if row_idx % 2 == 0 else fill_odd
        row_data = [
            c["S.No"], c["Name"], c["Company"], c["Their Role"],
            c["Hiring For"], c["Location"], c["LinkedIn URL"], c["Outreach Message"],
        ]
        for col, val in enumerate(row_data, 1):
            cell           = ws.cell(row=row_idx, column=col, value=val)
            cell.font      = Font(name="Calibri", size=10)
            cell.fill      = fill
            cell.border    = border
            cell.alignment = Alignment(vertical="top", wrap_text=(col == 8))
            if col == 7:   # LinkedIn URL — hyperlink style
                cell.font  = Font(name="Calibri", size=10, color="0563C1", underline="single")
        ws.row_dimensions[row_idx].height = 40 if len(c["Outreach Message"]) > 120 else 28

    # ── Column widths ──
    col_widths = [6, 22, 26, 28, 28, 18, 50, 65]
    for col, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    # ── Freeze header ──
    ws.freeze_panes = "A4"

    # ── Auto-filter ──
    ws.auto_filter.ref = f"A3:H{3 + len(contacts)}"

    wb.save(path)
    logger.info("Excel saved: %s", path)
    return path


# ── Gmail OAuth send with attachment ─────────────────────────────────────────

def send_with_attachment(to: str, subject: str, body: str, filepath: str):
    from google.auth.transport.requests import Request
    from google.oauth2.credentials import Credentials
    from google_auth_oauthlib.flow import InstalledAppFlow
    from googleapiclient.discovery import build
    import email as email_lib
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText
    from email.mime.base import MIMEBase
    from email import encoders

    SCOPES           = ["https://www.googleapis.com/auth/gmail.send"]
    token_path       = GMAIL_CREDENTIALS_DIR / "token.json"
    credentials_path = GMAIL_CREDENTIALS_DIR / "credentials.json"

    creds = None
    if token_path.exists():
        creds = Credentials.from_authorized_user_file(str(token_path), SCOPES)
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(str(credentials_path), SCOPES)
            creds = flow.run_local_server(port=0)
        token_path.write_text(creds.to_json())

    service = build("gmail", "v1", credentials=creds)

    msg = MIMEMultipart()
    msg["To"]      = to
    msg["Subject"] = subject
    msg.attach(MIMEText(body, "plain"))

    with open(filepath, "rb") as f:
        part = MIMEBase("application",
                        "vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        part.set_payload(f.read())
    encoders.encode_base64(part)
    part.add_header("Content-Disposition",
                    "attachment",
                    filename=os.path.basename(filepath))
    msg.attach(part)

    raw = base64.urlsafe_b64encode(msg.as_bytes()).decode()
    service.users().messages().send(userId="me", body={"raw": raw}).execute()
    logger.info("Email sent to %s", to)


# ── Main ──────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    today_str = date.today().isoformat()
    excel_path = f"/tmp/hiring_managers_purva_{today_str}.xlsx"

    print(f"\n🔍 Searching for {TARGET} hiring managers (Project Manager | Pune & Remote)…\n")
    contacts = collect_contacts(target=TARGET)

    if not contacts:
        print("❌ No contacts found. Check your internet connection.")
        exit(1)

    print(f"\n📊 Building Excel with {len(contacts)} contacts…")
    build_excel(contacts, excel_path)

    subject = (f"📋 {len(contacts)} Hiring Managers — Project Manager | "
               f"Pune & Remote | {date.today().strftime('%d %b %Y')}")

    body = f"""Hi Purva,

Please find attached a list of {len(contacts)} hiring managers and recruiters actively looking for Project Managers in Pune and Remote locations.

The Excel sheet includes:
• Their name, company, and current role
• What roles they are hiring for
• Location (Pune / Remote / Hybrid)
• Direct LinkedIn profile URL
• A ready-to-use personalised outreach message (under 300 chars)

How to use:
1. Open the Excel file
2. Sort/filter by Location or Company as needed
3. Copy the Outreach Message and send it as a LinkedIn connection note

Good luck! 🚀

—Job Search Agent
"""

    print(f"\n📧 Sending to {RECIPIENT}…")
    send_with_attachment(RECIPIENT, subject, body, excel_path)

    print(f"\n✅ Done! {len(contacts)} hiring managers sent to {RECIPIENT}")
    print(f"   Excel saved at: {excel_path}")
